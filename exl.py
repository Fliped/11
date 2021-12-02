
import openpyxl
import time

excel_01=openpyxl.load_workbook(r'D:\github\excel.xlsx')
#excel_01.sheetnames('Sheet1')
Sheet1=excel_01['Sheet1']
# She["C2"]="PASS"
for i in range (2,Sheet1.max_row+1):
    #print(Sheet1.cell(i,2).value)
    res='ada shl 42 ada sada   sadad'
    if str(Sheet1.cell(i,2).value) in res:
        Sheet1.cell(i,3).value="PASS"
    else:
        Sheet1.cell(i,3).value="Fail"

time.sleep(3)
excel_01.save('excel.xlsx')
#excel_01.get_sheet_by_name('Sheet1')